from __future__ import annotations

from pathlib import Path

from convert_to_md.context import ConvertContext
from convert_to_md.converters.base import BaseConverter
from convert_to_md.ir import DocumentIR


class XlsxConverter(BaseConverter):
    name = "xlsx"
    kinds = ("xlsx",)

    def convert(self, path: Path, ctx: ConvertContext) -> DocumentIR:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), data_only=True, read_only=True)
        doc = DocumentIR(source=path, title=path.stem)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            doc.heading(str(sheet_name), level=2)
            rows: list[list[str]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= ctx.max_table_rows:
                    break
                cells = ["" if c is None else str(c) for c in row]
                # skip fully empty rows
                if any(c.strip() for c in cells):
                    rows.append(cells)
            if rows:
                # trim trailing empty columns
                max_w = max(len(r) for r in rows)
                # also trim right-side empties across sheet
                last_col = 0
                for r in rows:
                    for j in range(len(r) - 1, -1, -1):
                        if r[j].strip():
                            last_col = max(last_col, j + 1)
                            break
                rows = [r[:last_col] + [""] * max(0, last_col - len(r)) for r in rows]
                doc.table(rows)
            else:
                doc.paragraph("_(empty sheet)_")

        wb.close()
        return doc
